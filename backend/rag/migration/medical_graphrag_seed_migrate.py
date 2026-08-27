from __future__ import annotations

import argparse
import json
import time
from pathlib import Path
from typing import Any

import requests


def _login(api_base: str, username: str, password: str) -> str:
    response = requests.post(
        f"{api_base.rstrip('/')}/auth/login",
        json={"username": username, "password": password},
        timeout=30,
    )
    response.raise_for_status()
    return response.json()["access_token"]


def _headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _upload(
    api_base: str,
    token_ref: dict[str, str],
    file_path: Path,
    *,
    username: str,
    password: str,
    source_name: str,
    source_type: str,
    source_tier: str,
    license_name: str,
    collection: str,
    dry_run: bool,
) -> dict[str, Any]:
    payload = {
        "source_name": source_name,
        "source_type": source_type,
        "source_tier": source_tier,
        "license": license_name,
        "collection": collection,
        "ingest_mode": "seed_migration",
        "description": "Created by medical_graphrag_seed_migrate.py",
        "create_ingest_run": "true",
    }
    if dry_run:
        return {"dry_run": True, "file": str(file_path), "payload": payload}
    with file_path.open("rb") as fh:
        response = _request_with_refresh(
            "POST",
            f"{api_base.rstrip('/')}/rag-admin/upload",
            token_ref=token_ref,
            username=username,
            password=password,
            api_base=api_base,
            data=payload,
            files={"file": (file_path.name, fh, "application/octet-stream")},
            timeout=120,
        )
    response.raise_for_status()
    return response.json()


def _start(
    api_base: str,
    token_ref: dict[str, str],
    run_id: str,
    dry_run: bool,
    *,
    username: str,
    password: str,
) -> dict[str, Any]:
    if dry_run:
        return {"dry_run": True, "would_start": run_id}
    response = _request_with_refresh(
        "POST",
        f"{api_base.rstrip('/')}/rag-admin/ingest-runs/{run_id}/start",
        token_ref=token_ref,
        username=username,
        password=password,
        api_base=api_base,
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


def _request_with_refresh(
    method: str,
    url: str,
    *,
    token_ref: dict[str, str],
    username: str,
    password: str,
    api_base: str,
    **kwargs,
) -> requests.Response:
    response = requests.request(method, url, headers=_headers(token_ref["token"]), **kwargs)
    if response.status_code == 401 and username and password:
        token_ref["token"] = _login(api_base, username, password)
        response = requests.request(method, url, headers=_headers(token_ref["token"]), **kwargs)
    return response


def _poll(
    api_base: str,
    token_ref: dict[str, str],
    run_id: str,
    *,
    username: str,
    password: str,
    timeout_seconds: int = 900,
) -> dict[str, Any]:
    deadline = time.time() + timeout_seconds
    last: dict[str, Any] = {}
    while time.time() < deadline:
        response = _request_with_refresh(
            "GET",
            f"{api_base.rstrip('/')}/rag-admin/ingest-runs/{run_id}",
            token_ref=token_ref,
            username=username,
            password=password,
            api_base=api_base,
            timeout=30,
        )
        response.raise_for_status()
        last = response.json()
        if last.get("status") in {"completed", "failed", "quarantined", "cancelled"}:
            return last
        time.sleep(5)
    return {**last, "poll_timeout": True}


def _latest_run_for_source(api_base: str, token_ref: dict[str, str], source_id: str, username: str, password: str) -> dict[str, Any] | None:
    response = _request_with_refresh(
        "GET",
        f"{api_base.rstrip('/')}/rag-admin/ingest-runs",
        token_ref=token_ref,
        username=username,
        password=password,
        api_base=api_base,
        params={"source_id": source_id, "limit": 1},
        timeout=30,
    )
    response.raise_for_status()
    items = response.json().get("items") or []
    return items[0] if items else None


def _create_run_for_source(
    api_base: str,
    token_ref: dict[str, str],
    source_id: str,
    *,
    username: str,
    password: str,
    dry_run: bool,
) -> dict[str, Any]:
    if dry_run:
        return {"dry_run": True, "would_create_run_for_source": source_id}
    response = _request_with_refresh(
        "POST",
        f"{api_base.rstrip('/')}/rag-admin/sources/{source_id}/ingest-runs",
        token_ref=token_ref,
        username=username,
        password=password,
        api_base=api_base,
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


def _mark_failed(
    api_base: str,
    token_ref: dict[str, str],
    run_id: str,
    *,
    username: str,
    password: str,
    reason: str,
) -> dict[str, Any]:
    response = _request_with_refresh(
        "POST",
        f"{api_base.rstrip('/')}/rag-admin/ingest-runs/{run_id}/mark-failed",
        token_ref=token_ref,
        username=username,
        password=password,
        api_base=api_base,
        data={"reason": reason},
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


def _stats(
    api_base: str,
    token_ref: dict[str, str],
    *,
    username: str,
    password: str,
    dry_run: bool,
) -> dict[str, Any] | None:
    if dry_run:
        return None
    response = _request_with_refresh(
        "GET",
        f"{api_base.rstrip('/')}/rag-admin/stats",
        token_ref=token_ref,
        username=username,
        password=password,
        api_base=api_base,
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


def _excel_data_row_count(file_path: Path) -> int:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        return max((sheet.max_row or 1) - 1, 0)
    finally:
        workbook.close()


def _make_drug_sample(drug_root: Path, row_limit: int, out_dir: Path, *, row_offset: int = 0) -> Path | None:
    import pandas as pd

    rows = []
    loaded = 0
    remaining_offset = max(row_offset, 0)
    for file_path in sorted(drug_root.glob("*.xlsx")):
        file_rows = _excel_data_row_count(file_path)
        if remaining_offset >= file_rows:
            remaining_offset -= file_rows
            continue
        remaining = max(row_limit - loaded, 0)
        if remaining <= 0:
            break
        local_skip = remaining_offset
        nrows = min(remaining, max(file_rows - local_skip, 0))
        remaining_offset = 0
        if nrows <= 0:
            continue
        frame = pd.read_excel(file_path, skiprows=range(1, local_skip + 1), nrows=nrows)
        if frame.empty:
            continue
        rows.append(frame)
        loaded += len(frame)
        if loaded >= row_limit:
            break
    if not rows:
        return None
    sample = pd.concat(rows, ignore_index=True).head(row_limit)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"drug_label_sample_offset_{row_offset}_limit_{len(sample)}.xlsx"
    sample.to_excel(out_path, index=False)
    return out_path


def run() -> int:
    parser = argparse.ArgumentParser(description="Seed-migrate small local RAG samples into medical-graphrag via admin APIs.")
    parser.add_argument("--api-base", default="http://localhost:8026/api/v1")
    parser.add_argument("--username", default="admin")
    parser.add_argument("--password", default="admin123")
    parser.add_argument("--token", default="")
    parser.add_argument("--pdf-root", default="backend/data_PDF")
    parser.add_argument("--drug-root", default="backend/drug_data")
    parser.add_argument("--pdf-limit", type=int, default=10)
    parser.add_argument("--drug-row-offset", type=int, default=0)
    parser.add_argument("--drug-row-limit", type=int, default=500)
    parser.add_argument("--drug-batch-rows", type=int, default=0)
    parser.add_argument("--start", action="store_true")
    parser.add_argument("--poll", action="store_true")
    parser.add_argument("--rerun-duplicates", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--out", default="rag/reports/medical_graphrag_migration/seed_migration_report.json")
    args = parser.parse_args()

    root = Path.cwd()
    out_path = (root / args.out).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    token = args.token or ("dry_run_token" if args.dry_run else _login(args.api_base, args.username, args.password))
    token_ref = {"token": token}
    auth_username = args.username if not args.token else ""
    auth_password = args.password if not args.token else ""
    stats_before = _stats(
        args.api_base,
        token_ref,
        username=auth_username,
        password=auth_password,
        dry_run=args.dry_run,
    )

    entries: list[dict[str, Any]] = []
    pdf_root = (root / args.pdf_root).resolve()
    for pdf_path in sorted(pdf_root.rglob("*.pdf"))[: args.pdf_limit]:
        upload = _upload(
            args.api_base,
            token_ref,
            pdf_path,
            username=auth_username,
            password=auth_password,
            source_name=pdf_path.stem,
            source_type="guideline_pdf",
            source_tier="T1",
            license_name="local_review_required",
            collection="medical_guideline_v2",
            dry_run=args.dry_run,
        )
        entry = {"kind": "pdf", "file": str(pdf_path), "upload": upload}
        run_id = (upload.get("ingest_run") or {}).get("ingest_run_id")
        if args.start and run_id and not upload.get("duplicate"):
            entry["start"] = _start(
                args.api_base,
                token_ref,
                run_id,
                args.dry_run,
                username=auth_username,
                password=auth_password,
            )
            if args.poll and not args.dry_run:
                entry["final"] = _poll(
                    args.api_base,
                    token_ref,
                    run_id,
                    username=auth_username,
                    password=auth_password,
                )
        elif upload.get("duplicate") and upload.get("source", {}).get("source_id"):
            source_id = upload["source"]["source_id"]
            latest = _latest_run_for_source(args.api_base, token_ref, source_id, auth_username, auth_password)
            entry["existing_run"] = latest
            if args.rerun_duplicates:
                created = _create_run_for_source(
                    args.api_base,
                    token_ref,
                    source_id,
                    username=auth_username,
                    password=auth_password,
                    dry_run=args.dry_run,
                )
                entry["created_run_for_duplicate"] = created
                run_id = created.get("ingest_run_id")
                if args.start and run_id:
                    entry["start"] = _start(
                        args.api_base,
                        token_ref,
                        run_id,
                        args.dry_run,
                        username=auth_username,
                        password=auth_password,
                    )
                if args.poll and run_id and not args.dry_run:
                    final = _poll(
                        args.api_base,
                        token_ref,
                        run_id,
                        username=auth_username,
                        password=auth_password,
                    )
                    entry["final"] = final
                    if final.get("poll_timeout") and final.get("status") in {"queued", "processing"}:
                        entry["marked_failed"] = _mark_failed(
                            args.api_base,
                            token_ref,
                            run_id,
                            username=auth_username,
                            password=auth_password,
                            reason="seed_migration_poll_timeout",
                        )
            elif args.poll and latest:
                entry["final"] = _poll(
                    args.api_base,
                    token_ref,
                    latest["ingest_run_id"],
                    username=auth_username,
                    password=auth_password,
                )
        entries.append(entry)

    drug_samples: list[tuple[int, int, Path]] = []
    drug_root = (root / args.drug_root).resolve()
    if args.drug_batch_rows and args.drug_batch_rows > 0:
        end_offset = args.drug_row_offset + args.drug_row_limit
        offset = args.drug_row_offset
        while offset < end_offset:
            limit = min(args.drug_batch_rows, end_offset - offset)
            sample = _make_drug_sample(drug_root, limit, out_path.parent, row_offset=offset)
            if sample is not None:
                drug_samples.append((offset, limit, sample))
            offset += limit
    else:
        sample = _make_drug_sample(drug_root, args.drug_row_limit, out_path.parent, row_offset=args.drug_row_offset)
        if sample is not None:
            drug_samples.append((args.drug_row_offset, args.drug_row_limit, sample))

    for drug_offset, drug_limit, drug_sample in drug_samples:
        upload = _upload(
            args.api_base,
            token_ref,
            drug_sample,
            username=auth_username,
            password=auth_password,
            source_name=f"CFDA/NMPA local drug sample offset {drug_offset} limit {drug_limit}",
            source_type="drug_excel",
            source_tier="T1",
            license_name="local_official_snapshot_review_required",
            collection="drug_label_v2",
            dry_run=args.dry_run,
        )
        entry = {
            "kind": "drug_excel",
            "file": str(drug_sample),
            "drug_row_offset": drug_offset,
            "drug_row_limit": drug_limit,
            "upload": upload,
        }
        run_id = (upload.get("ingest_run") or {}).get("ingest_run_id")
        if args.start and run_id and not upload.get("duplicate"):
            entry["start"] = _start(
                args.api_base,
                token_ref,
                run_id,
                args.dry_run,
                username=auth_username,
                password=auth_password,
            )
            if args.poll and not args.dry_run:
                entry["final"] = _poll(
                    args.api_base,
                    token_ref,
                    run_id,
                    username=auth_username,
                    password=auth_password,
                    timeout_seconds=1800,
                )
        elif upload.get("duplicate") and upload.get("source", {}).get("source_id"):
            source_id = upload["source"]["source_id"]
            latest = _latest_run_for_source(args.api_base, token_ref, source_id, auth_username, auth_password)
            entry["existing_run"] = latest
            if args.rerun_duplicates:
                created = _create_run_for_source(
                    args.api_base,
                    token_ref,
                    source_id,
                    username=auth_username,
                    password=auth_password,
                    dry_run=args.dry_run,
                )
                entry["created_run_for_duplicate"] = created
                run_id = created.get("ingest_run_id")
                if args.start and run_id:
                    entry["start"] = _start(
                        args.api_base,
                        token_ref,
                        run_id,
                        args.dry_run,
                        username=auth_username,
                        password=auth_password,
                    )
                if args.poll and run_id and not args.dry_run:
                    final = _poll(
                        args.api_base,
                        token_ref,
                        run_id,
                        username=auth_username,
                        password=auth_password,
                        timeout_seconds=1800,
                    )
                    entry["final"] = final
                    if final.get("poll_timeout") and final.get("status") in {"queued", "processing"}:
                        entry["marked_failed"] = _mark_failed(
                            args.api_base,
                            token_ref,
                            run_id,
                            username=auth_username,
                            password=auth_password,
                            reason="seed_migration_poll_timeout",
                        )
            elif args.poll and latest:
                entry["final"] = _poll(
                    args.api_base,
                    token_ref,
                    latest["ingest_run_id"],
                    username=auth_username,
                    password=auth_password,
                    timeout_seconds=1800,
                )
        entries.append(entry)

    stats_after = _stats(
        args.api_base,
        token_ref,
        username=auth_username,
        password=auth_password,
        dry_run=args.dry_run,
    )
    report = {
        "api_base": args.api_base,
        "dry_run": args.dry_run,
        "stats_before": stats_before,
        "stats_after": stats_after,
        "entries": entries,
    }
    out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(str(out_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(run())
